import os
import json
import re
import glob
from datetime import datetime
from flask import Flask, request, jsonify, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
APPLICATIONS_DIR = os.path.join(BASE_DIR, 'applications')
STATIC_DIR = os.path.join(BASE_DIR, 'static')

app = Flask(__name__, static_folder=STATIC_DIR)

os.makedirs(APPLICATIONS_DIR, exist_ok=True)


def sanitize_company_name(name):
    sanitized = re.sub(r'[^a-zA-Z0-9]', '-', name)
    sanitized = re.sub(r'-+', '-', sanitized)
    sanitized = sanitized.strip('-')
    return sanitized


def generate_filename(agency, applications_dir=None):
    if applications_dir is None:
        applications_dir = APPLICATIONS_DIR
    today = datetime.now().strftime('%Y-%m-%d')
    pattern = os.path.join(applications_dir, f"REQ-{today}-*.md")
    existing = glob.glob(pattern)
    max_seq = 0
    for f in existing:
        basename = os.path.basename(f)
        match = re.match(r'REQ-\d{4}-\d{2}-\d{2}-(\d{3})-', basename)
        if match:
            seq = int(match.group(1))
            if seq > max_seq:
                max_seq = seq
    next_seq = f"{max_seq + 1:03d}"
    company = sanitize_company_name(agency)
    return f"REQ-{today}-{next_seq}-{company}.md"


def generate_markdown(data, request_id):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    req_type = "New Laptop" if data.get("requestType") == "new" else "Replacement Laptop"
    unit_cost = data.get("unitCost", 0)
    quantity = data.get("quantity", 1)
    total_cost = unit_cost * quantity
    currency = data.get("currency", "USD")

    def fmt_cost(val):
        return f"{val:,.0f}" if val == int(val) else f"{val:,.2f}"

    def checkbox(val):
        return "[x]" if val else "[ ]"

    lines = []
    lines.append(f"# Laptop Request — {request_id}")
    lines.append("")
    lines.append(f"**Generated:** {now}")
    lines.append(f"**Request ID:** {request_id}")
    lines.append("")
    lines.append("## Request Type")
    lines.append(f"- Type: {req_type}")
    lines.append("")
    lines.append("## Requestor Information")
    lines.append(f"- Name: {data.get('requestorName', '')}")
    lines.append(f"- Agency/Company: {data.get('agency', '')}")
    lines.append(f"- Market: {data.get('market', '')}")
    lines.append(f"- Email: {data.get('email', '')}")
    lines.append(f"- Staff Category: {data.get('staffCategory', 'Standard')}")
    lines.append("")
    lines.append("## Stock Verification")
    lines.append(f"- {checkbox(data.get('stockNoBuffer'))} No buffer or almost-new laptops in existing stock")
    lines.append(f"- {checkbox(data.get('stockListingUpToDate'))} 2026 Stock listing file is up to date")
    lines.append("")
    lines.append("## Laptop Specifications")
    lines.append(f"- EUC Persona: {data.get('eucPersona', '')}")
    lines.append(f"- EUC Standards Version: {data.get('eucStandardsVersion', '2025/Q4')}")
    device_type = data.get('deviceType', '')
    if device_type:
        lines.append(f"- Device Type: {device_type.capitalize()}")
    lines.append(f"- Model: {data.get('laptopModel', '')}")
    make = data.get('make', '')
    if make:
        lines.append(f"- Make: {make}")
    mpn = data.get('mpn', '')
    if mpn:
        lines.append(f"- MPN: {mpn}")
    lines.append(f"- OS: {data.get('os', '')}")
    mac_just = data.get('macOsJustification')
    lines.append(f"- macOS Justification: {mac_just if mac_just else 'N/A'}")
    lines.append(f"- Quantity: {quantity}")
    lines.append(f"- Specs: {data.get('specs', '')}")
    lines.append("")
    lines.append("## Cost & Sourcing")
    lines.append(f"- Unit Cost: {fmt_cost(unit_cost)} {currency}")
    lines.append(f"- Total Cost: {fmt_cost(total_cost)} {currency}")
    is_apple = data.get('os') == 'macOS' or (data.get('make', '') or '').lower() == 'apple'
    cost_source = "Cost from WPP designated supplier (Computacenter / Apple Direct)" if is_apple else "Cost from Dell Direct portal / approved partner"
    lines.append(f"- {checkbox(data.get('costFromDell'))} {cost_source}")
    lines.append(f"- {checkbox(data.get('costExcludesTax'))} Cost excludes local taxes")
    lines.append("")

    if data.get("requestType") == "new":
        lines.append("## New Hire Details")
        lines.append(f"- Number of New Hires: {data.get('newHireCount', '')}")
        lines.append(f"- Expected Join Date: {data.get('joinDate', '')}")
        persona = data.get('newHirePersona')
        lines.append(f"- EUC Persona Override: {persona if persona else 'N/A'}")
        lines.append(f"- Available Functional Laptops: {data.get('availableLaptops', '')}")
        lines.append("")
    elif data.get("requestType") == "replacement":
        lines.append("## Current Device Information")
        dev_make = data.get('currentDeviceMake', '')
        dev_model = data.get('currentDeviceModel', '')
        lines.append(f"- Current Device: {dev_make} {dev_model}")
        serial = data.get('currentSerialNumber')
        if serial:
            lines.append(f"- Serial Number: {serial}")
        dev_age = data.get('currentDeviceAge')
        if dev_age is not None:
            lines.append(f"- Device Age: {dev_age} years")
        dev_specs = data.get('currentDeviceSpecs')
        if dev_specs:
            lines.append(f"- Current Device Specs: {dev_specs}")
        lines.append("")

        lines.append("## Replacement Reason")
        lines.append(f"- Reason: {data.get('currentCondition', '')}")
        lines.append(f"- Issue Details: {data.get('diagnostics', '')}")
        lines.append(f"- {checkbox(data.get('eusConfirmed'))} EUS / IT Support confirmed unfixable")
        lines.append("")

        lines.append("## Current Workaround")
        workaround = data.get('currentWorkaround', '')
        lines.append(f"- Workaround: {workaround}")
        wa_details = data.get('workaroundDetails')
        if wa_details:
            lines.append(f"- Details: {wa_details}")
        lines.append("")

        lines.append("## Additional Justification")
        stock_just = data.get('stockNotUsedJustification')
        if stock_just:
            lines.append(f"- Why Existing Stock Not Used: {stock_just}")
        non_std = data.get('nonStandardJustification')
        if non_std:
            lines.append(f"- Non-Standard Config Justification: {non_std}")
        if not stock_just and not non_std:
            lines.append("(none)")
        lines.append("")

    lines.append("## Additional Comments")
    comments = data.get("comments", "").strip()
    lines.append(comments if comments else "(none)")
    lines.append("")

    return "\n".join(lines)


def generate_excel(data, request_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "APAC Procurement Tracker"

    headers = [
        "Requestor (EUS LEAD) Email ID",       # A
        "Requestor (BU) Email ID",              # B
        "Agency",                               # C
        "Date requested",                       # D
        "Date Required",                        # E
        "Market",                               # F
        "Device Type\n(Windows or Mac)",        # G
        "Model\n* PC STD is now Dell Only",     # H
        "Quantity",                              # I
        "Local Currency",                       # J
        "Cost per Device (Local ccy)",          # K
        "Total Cost (Local Ccy)",               # L
        "Cost per Device (USD)",                # M
        "Total Cost (USD)",                     # N
        "Rationale\n(Separate entry for New Joiner or replacement)",  # O
        "Using existing stock or new purchase", # P
        "SS SC Review Comments\nDate Approved by Regional TechOps Director?",  # Q
        "ET Legal Entity Presence in Location?",# R
        "Lead Entity in Market",                # S
        "BFC Code",                             # T
    ]

    header_fill = PatternFill(start_color="0033A0", end_color="0033A0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 14)

    # Derive device type for tracker
    os_val = data.get('os', '')
    device_type_str = "Mac" if os_val == "macOS" else "Windows"

    # Build rationale (column O)
    quantity = data.get('quantity', 1)
    unit_cost = data.get('unitCost', 0)

    if data.get('requestType') == 'new':
        rationale = "New Joiner - {} new hires joining {}. {}".format(
            data.get('newHireCount', ''),
            data.get('joinDate', ''),
            data.get('availableLaptops', '')
        )
    else:
        rationale = "{} - {} {} ({}yr). Workaround: {}".format(
            data.get('currentCondition', ''),
            data.get('currentDeviceMake', ''),
            data.get('currentDeviceModel', ''),
            data.get('currentDeviceAge', ''),
            data.get('currentWorkaround', '')
        )

    # Build comments in list format
    comments = data.get('comments', '') or ''
    comment_lines = []

    if data.get('requestType') == 'new':
        comment_lines.append(f"- Request Type: New Laptop")
        comment_lines.append(f"- Persona: {data.get('eucPersona', '')}")
        comment_lines.append(f"- New Hires: {data.get('newHireCount', '')} joining {data.get('joinDate', '')}")
        comment_lines.append(f"- Available Laptops: {data.get('availableLaptops', '')}")
        if data.get('os') == 'macOS' and data.get('macOsJustification'):
            comment_lines.append(f"- macOS Justification: {data.get('macOsJustification', '')}")
        if comments:
            comment_lines.append(f"- Additional Notes: {comments}")
    else:
        comment_lines.append(f"- Request Type: Replacement")
        comment_lines.append(f"- Persona: {data.get('eucPersona', '')}")
        comment_lines.append(f"- Current Device: {data.get('currentDeviceMake', '')} {data.get('currentDeviceModel', '')}")
        if data.get('currentSerialNumber'):
            comment_lines.append(f"- Serial Number: {data.get('currentSerialNumber')}")
        comment_lines.append(f"- Device Age: {data.get('currentDeviceAge', '')} years")
        if data.get('currentDeviceSpecs'):
            comment_lines.append(f"- Current Specs: {data.get('currentDeviceSpecs')}")
        comment_lines.append(f"- Reason: {data.get('currentCondition', '')}")
        comment_lines.append(f"- Issue Details: {data.get('diagnostics', '')}")
        eus_str = "Yes" if data.get('eusConfirmed') else "No"
        comment_lines.append(f"- EUS/IT Confirmed Unfixable: {eus_str}")
        comment_lines.append(f"- Current Workaround: {data.get('currentWorkaround', '')}")
        if data.get('workaroundDetails'):
            comment_lines.append(f"- Workaround Details: {data.get('workaroundDetails')}")
        if data.get('os') == 'macOS' and data.get('macOsJustification'):
            comment_lines.append(f"- macOS Justification: {data.get('macOsJustification', '')}")
        if data.get('stockNotUsedJustification'):
            comment_lines.append(f"- Why Stock Not Used: {data.get('stockNotUsedJustification')}")
        if comments:
            comment_lines.append(f"- Additional Notes: {comments}")

    detail_comment = "\n".join(comment_lines)

    # Local cost calculations
    local_cost_per = data.get('localCostPerDevice', '') or ''
    local_total = ''
    if local_cost_per and isinstance(local_cost_per, (int, float)):
        local_total = local_cost_per * quantity

    today = datetime.now().strftime('%Y-%m-%d')

    row_data = [
        data.get('eusLeadEmail', ''),                   # A
        data.get('buEmail', '') or '',                   # B
        data.get('agency', ''),                          # C
        today,                                           # D
        data.get('dateRequired', ''),                    # E
        data.get('market', ''),                          # F
        device_type_str,                                 # G
        data.get('laptopModel', ''),                     # H
        quantity,                                        # I
        data.get('localCurrency', '') or '',             # J
        local_cost_per,                                  # K
        local_total,                                     # L
        unit_cost,                                       # M
        unit_cost * quantity,                            # N
        rationale,                                       # O
        data.get('stockOrNewPurchase', ''),              # P
        '',                                              # Q - leave empty
        data.get('etLegalEntity', '') or '',             # R
        data.get('leadEntityInMarket', '') or '',        # S
        data.get('bfcCode', '') or '',                   # T
    ]

    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=2, column=col_idx, value=value)

    # Comments column — matches original tracker column W (col 23)
    # Headers U/V are approval columns left empty, W=23 is Comments
    ws.cell(row=1, column=23, value="Comments")
    ws.cell(row=1, column=23).fill = header_fill
    ws.cell(row=1, column=23).font = header_font
    ws.cell(row=1, column=23).alignment = Alignment(wrap_text=True)
    ws.column_dimensions['W'].width = 60
    ws.cell(row=2, column=23, value=detail_comment)
    ws.cell(row=2, column=23).alignment = Alignment(wrap_text=True)

    filepath = os.path.join(APPLICATIONS_DIR, request_id + '.xlsx')
    wb.save(filepath)
    return request_id + '.xlsx'


def validate_request(data):
    errors = []
    for field in ['requestorName', 'agency', 'market', 'staffCategory']:
        if not data.get(field, '').strip():
            errors.append(f"{field} is required")

    email = data.get('email', '')
    if not re.match(r'^[^@]+@[^@]+\.[^@]+$', email):
        errors.append("email format invalid")

    if data.get('requestType') not in ('new', 'replacement'):
        errors.append("requestType must be 'new' or 'replacement'")

    if not data.get('stockNoBuffer'):
        errors.append("stockNoBuffer confirmation is required")
    if not data.get('stockListingUpToDate'):
        errors.append("stockListingUpToDate confirmation is required")

    if not data.get('eucPersona', '').strip():
        errors.append("eucPersona is required")
    if not data.get('laptopModel', '').strip():
        errors.append("laptopModel is required")
    if data.get('os') not in ('Windows', 'macOS'):
        errors.append("os must be 'Windows' or 'macOS'")
    # macOS justification is optional

    # deviceType is optional for backward compatibility
    if data.get('deviceType') and data.get('deviceType') not in ('laptop', 'desktop'):
        errors.append("deviceType must be 'laptop' or 'desktop'")

    if data.get('eucPersona') == 'Other':
        if not data.get('nonStandardJustification') or not data.get('nonStandardJustification', '').strip():
            errors.append("nonStandardJustification is required when EUC Persona is Other")

    quantity = data.get('quantity')
    if not isinstance(quantity, (int, float)) or quantity < 1:
        errors.append("quantity must be a positive number")
    if not data.get('specs', '').strip():
        errors.append("specs is required")

    unit_cost = data.get('unitCost')
    if not isinstance(unit_cost, (int, float)) or unit_cost <= 0:
        errors.append("unitCost must be a positive number")
    if not data.get('currency', '').strip():
        errors.append("currency is required")
    if not data.get('costFromDell'):
        errors.append("costFromDell confirmation is required")
    if not data.get('costExcludesTax'):
        errors.append("costExcludesTax confirmation is required")

    if data.get('requestType') == 'new':
        if not isinstance(data.get('newHireCount'), (int, float)) or data.get('newHireCount', 0) < 1:
            errors.append("newHireCount is required for new laptop requests")
        if not data.get('joinDate') or not str(data.get('joinDate', '')).strip():
            errors.append("joinDate is required for new laptop requests")
        if not data.get('availableLaptops') or not str(data.get('availableLaptops', '')).strip():
            errors.append("availableLaptops is required for new laptop requests")

    if data.get('requestType') == 'replacement':
        if not data.get('currentCondition') or not str(data.get('currentCondition', '')).strip():
            errors.append("currentCondition is required for replacement requests")
        if not data.get('diagnostics') or not str(data.get('diagnostics', '')).strip():
            errors.append("diagnostics is required for replacement requests")
        if not data.get('eusConfirmed'):
            errors.append("eusConfirmed is required for replacement requests")

    # New fields - only validate if present in data (backward compatible)
    if 'eusLeadEmail' in data:
        eus_email = data.get('eusLeadEmail', '')
        if not eus_email or not re.match(r'^[^@]+@[^@]+\.[^@]+$', eus_email):
            errors.append("eusLeadEmail format invalid")
    if 'dateRequired' in data:
        if not data.get('dateRequired', '').strip():
            errors.append("dateRequired is required")
    if 'bfcCode' in data:
        if not data.get('bfcCode', '').strip():
            errors.append("bfcCode is required")
    if 'stockOrNewPurchase' in data:
        if not data.get('stockOrNewPurchase', '').strip():
            errors.append("stockOrNewPurchase is required")
    if 'etLegalEntity' in data:
        if not data.get('etLegalEntity', '').strip():
            errors.append("etLegalEntity is required")

    return errors


@app.route('/')
def index():
    return send_from_directory(STATIC_DIR, 'index.html')


@app.route('/api/health')
def health():
    return jsonify({'status': 'ok'})


@app.route('/api/save', methods=['POST'])
def save_request():
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'error': 'No data provided', 'details': []}), 400

    errors = validate_request(data)
    if errors:
        return jsonify({'success': False, 'error': 'Validation failed', 'details': errors}), 400

    filename = generate_filename(data.get('agency', 'Unknown'))
    request_id = filename.replace('.md', '')
    md_content = generate_markdown(data, request_id)

    filepath = os.path.join(APPLICATIONS_DIR, filename)
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(md_content)

    # Generate Excel file
    excel_filename = generate_excel(data, request_id)

    return jsonify({
        'success': True,
        'filename': filename,
        'excelFilename': excel_filename,
        'path': f'applications/{filename}'
    })


@app.route('/applications/<path:filename>')
def download_file(filename):
    return send_from_directory(APPLICATIONS_DIR, filename, as_attachment=True)


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8900, debug=True)
